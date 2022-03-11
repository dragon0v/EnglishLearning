# -*- coding: utf-8 -*-
"""
Created on Mon Aug  3 21:01:58 2020

@author: NeoBanana
"""

#word_tester


from tkinter import *
import tkinter.messagebox
import tkinter.font as tf
import time
import datetime
import os
import random
import webbrowser
from openpyxl import load_workbook
#from playsound_my import playsound
#from pydub import AudioSegment
#from pydub.playback import play
#import Play_mp3


#OK play sound
#OK play_sound改成异步操作，现在播放声音时其他操作都不行 -- 修改play_sound的block就行
#OK 退出询问是否保存
# 更改后将list内容也改咯
#TODO 记忆方法，词组备注，例句在正确输入后再出现，除非是一路enter过去的，增加is_shown_2?
#TODO visible & editable changelog  直接用text显示就行？最好用list
#TODO display how many words tested, etc
#TODO search word
#TODO? 不重复提问



#以下是全局变量
print(os.environ['USERNAME']) #可能不成功因为与环境变量有关
USERNAME = os.environ['USERNAME']
if USERNAME=="ToxicNeoBanana":
    PATH_SHANBAY = r"F:\OneDrive\shanbay1.xlsx"
    PATH_SHANBAY_BACKUP = os.path.join(os.path.expanduser('~'),"Desktop","shanbay1-%d.xlsx"%int(time.time()))
    PATH_AUDIO = r"D:\python\英语学习\voices"
elif USERNAME=="NeoBanana":
    PATH_SHANBAY = r"C:\Users\NeoBanana\OneDrive\shanbay1.xlsx"
    PATH_SHANBAY_BACKUP = os.path.join(os.path.expanduser('~'),"Desktop","shanbay1-%d.xlsx"%int(time.time()))
    PATH_AUDIO = r"D:\python\英语学习\voices"
else:
    PATH_SHANBAY = r"shanbay1.xlsx"
    PATH_SHANBAY_BACKUP = os.path.join(os.path.expanduser('~'),"Desktop","shanbay1-%d.xlsx"%int(time.time()))
    PATH_AUDIO = r"voices"



COLOR_BG_I = '#ffafaf'
COLOR_FG_I = '#ff2727'
COLOR_BG_P = '#c0e399'
COLOR_FG_P = '#548235'
COLOR_BG_U = '#acb9ca'
COLOR_FG_U = '#333f4f'
COLOR_BG_N = '#F7F762'
COLOR_FG_N = 'black'
COLOR_WRONG_WORD = 'red'
COLOR_PASS = '#444' #回车过去的单词颜色

class Tester():
    def __init__(self):
        self.win=Tk()
        self.win.geometry('480x512+800+380')
        self.win.title("word_tester")
        
        self.place_widgets()
        
        self.load_excel()
        
        self.testedlog=[] # 测试的单词记录，目前在prev模块使log功能无效
        self.changelog=[] # save all changes
        self.is_shown = False
        
        self.nextword()
        
        #此处可以用于debug print
        
        
        self.win.protocol("WM_DELETE_WINDOW",self.closewin)#按关闭键执行self.closewin
        
        self.win.mainloop()
    
    def place_widgets(self):
        self.canvas = Canvas(self.win,width=900,height=600,background='lightgrey')
        self.canvas.pack()
        
        self.word_entry = StringVar()
        ft = tf.Font(family='mv boli', size='14')
        self.entry_word = Entry(self.canvas,bg='yellow',textvariable=self.word_entry,font=ft)
        self.entry_word.place(x=10,y=10,width=160,height=26,anchor=NW)
        self.entry_word.bind("<Return>",self.process_enter)
        self.button_next = Button(self.canvas,text="next",command=lambda:self.process_next())
        self.button_next.place(x=180,y=10,width=60,height=50,anchor=NW)
        self.button_prev = Button(self.canvas,text="prev",command=lambda:self.process_prev()) # TODO 未完成
        self.button_prev.place(x=180,y=64,width=60,height=12,anchor=NW)
        self.button_repeat = Button(self.canvas,text="repeat",command=lambda:self.process_repeat())
        self.button_repeat.place(x=250,y=10,width=60,height=50,anchor=NW)
        self.button_show = Button(self.canvas,text="show",command=lambda:self.process_show())
        self.button_show.place(x=320,y=10,width=60,height=50,anchor=NW)
        self.button_baidu = Button(self.canvas,text="bing",command=lambda:self.open_url('bing'))
        self.button_baidu.place(x=390,y=10,width=60,height=24,anchor=NW)
        self.button_youdao = Button(self.canvas,text="youdao",command=lambda:self.open_url('youdao'))
        self.button_youdao.place(x=390,y=36,width=60,height=24,anchor=NW)
        self.button_save = Button(self.canvas,text="save",command=lambda:self.process_save())
        self.button_save.place(x=460,y=10,width=12,height=12,anchor=NW)
        
        self.button_ipu_i = Button(self.canvas,text="",command=lambda:self.set_ipu('i'),background=COLOR_BG_I)
        self.button_ipu_i.place(x=10,y=40,width=50,height=20,anchor=NW)
        self.button_ipu_p = Button(self.canvas,text="",command=lambda:self.set_ipu('p'),background=COLOR_BG_P)
        self.button_ipu_p.place(x=65,y=40,width=50,height=20,anchor=NW)
        self.button_ipu_u = Button(self.canvas,text="",command=lambda:self.set_ipu('u'),background=COLOR_BG_U)
        self.button_ipu_u.place(x=120,y=40,width=50,height=20,anchor=NW)
        
        self.label_paraphrase = Label(self.canvas,text="paraphrase")
        self.label_paraphrase.place(x=10,y=70)
        self.listbox_paraphrase = Listbox(self.canvas)
        self.listbox_paraphrase.place(x=10,y=90,width=230,height=200,anchor=NW)
        self.label_jyff = Label(self.canvas,text="记忆方法")
        self.label_jyff.place(x=250,y=70)
        self.button_addjyff = Button(self.canvas,text="add",command=lambda:self.add('jyff',self.text_jyff.get(1.0,100.0)[:-1])) #:-1去掉最后的换行
        self.button_addjyff.place(x=410,y=72,width=40,height=20,anchor=NW)
        self.text_jyff = Text(self.canvas)
        self.text_jyff.place(x=250,y=90,width=200,height=90,anchor=NW)
        self.label_czbz = Label(self.canvas,text="词组备注")
        self.label_czbz.place(x=250,y=180)
        self.button_addczbz = Button(self.canvas,text="add",command=lambda:self.add('czbz',self.text_czbz.get(1.0,100.0)[:-1]))
        self.button_addczbz.place(x=410,y=182,width=40,height=20,anchor=NW)
        self.text_czbz = Text(self.canvas)
        self.text_czbz.place(x=250,y=200,width=200,height=90,anchor=NW)
        
        
        self.label_example = Label(self.canvas,text="example")
        self.label_example.place(x=10,y=300)
        self.listbox_example = Listbox(self.canvas)
        self.listbox_example.place(x=10,y=320,width=440,height=104,anchor=NW)
        
        self.text_addexample = Text(self.canvas)
        self.text_addexample.place(x=10,y=434,width=400,height=72,anchor=NW)
        self.button_addexample = Button(self.canvas,text="add",command=lambda:self.add('example',self.text_addexample.get(1.0,100.0)[:-1]))
        self.button_addexample.place(x=410,y=434,width=40,height=20,anchor=NW)
        
    def load_excel(self):
        workbook = load_workbook(PATH_SHANBAY)#找到excel文件
        sheet = workbook["Sheet1"]#找到当前表格
        self.row_number = sheet.max_row
        print(self.row_number)
        
        self.list_ipu = []
        self.list_word = []
        list_paraphrase_raw = [] #need further process
        self.list_jyff = []
        self.list_czbz = []
        list_example_raw = [] #
        
        for i in range(2,self.row_number+1):
            self.list_ipu.append(sheet.cell(row=i,column=1).value)
            self.list_word.append(sheet.cell(row=i,column=2).value)
            list_paraphrase_raw.append(sheet.cell(row=i,column=3).value)
            self.list_jyff.append(sheet.cell(row=i,column=4).value)
            self.list_czbz.append(sheet.cell(row=i,column=5).value)
            list_example_raw.append(sheet.cell(row=i,column=6).value)
        
        #print(self.list_czbz[:10])
        
        self.list_paraphrase = self.parse_paraphrase(list_paraphrase_raw) # list in list
        self.list_example = self.parse_paraphrase(list_example_raw) # 
        #print(self.list_paraphrase[:10])
            
    def parse_paraphrase(self,list_paraphrase):
        #="n. 耙"&CHAR(10)&"vt. 耙地；使苦恼"&CHAR(10)&"vi. 被耙松"&CHAR(10)&"n. (Harrow)人名；(英)哈罗"
        after = []
        for each in list_paraphrase:
            if each == None:
                after.append([])
            else:
                p1 = each[2:-1]
                p2 = p1.split('"&CHAR(10)&"')
                after.append(p2) # list in list
        return after
    
    def play_sound(self,path):
        # 因为playsound库在中文路径报错，在这里尝试修复
        # 发现如果音频不存在也会报UnicodeDecodeError
        # contents below is copied from 'playsound'
        class PlaysoundException(Exception):
            pass
        
        def _playsoundWin(sound, block = True):
            '''
            Utilizes windll.winmm. Tested and known to work with MP3 and WAVE on
            Windows 7 with Python 2.7. Probably works with more file formats.
            Probably works on Windows XP thru Windows 10. Probably works with all
            versions of Python.
        
            Inspired by (but not copied from) Michael Gundlach <gundlach@gmail.com>'s mp3play:
            https://github.com/michaelgundlach/mp3play
        
            I never would have tried using windll.winmm without seeing his code.
            '''
            from ctypes import c_buffer, windll
            from random import random
            from time   import sleep
            # from sys    import getfilesystemencoding
        
            def winCommand(*command):
                buf = c_buffer(255)
#                command = ' '.join(command).encode(getfilesystemencoding())
#                win默认gbk作为系统内部编码，在命令行cosole调试时使用的是内部编码
                command = ' '.join(command).encode('gbk')
                errorCode = int(windll.winmm.mciSendStringA(command, buf, 254, 0))
                if errorCode:
                    errorBuffer = c_buffer(255)
                    windll.winmm.mciGetErrorStringA(errorCode, errorBuffer, 254)
                    exceptionMessage = ('\n    Error ' + str(errorCode) + ' for command:'
                                        '\n        ' + command.decode() +
                                        '\n    ' + errorBuffer.value.decode())
                    raise PlaysoundException(exceptionMessage)
                return buf.value
        
            alias = 'playsound_' + str(random())
            winCommand('open "' + sound + '" alias', alias)
            winCommand('set', alias, 'time format milliseconds')
            durationInMS = winCommand('status', alias, 'length')
            winCommand('play', alias, 'from 0 to', durationInMS.decode())
            
            # True则播放音频时不能进行其他操作，False则其他操作不受影响
            if block:
                sleep(float(durationInMS) / 1000.0)
        
        _playsoundWin(path,False)
    
    def nextword(self,designate=None):
        '''
        clear exsiting entries;
        show next word
        
        Parameters
        ----------
        designate : str, optional, use when processing 'prev'
            DESCRIPTION. designate a word to be next

        Returns
        -------
        None.

        '''
        self.is_shown = False
        self.word_entry.set('')
        self.listbox_paraphrase.delete(0,9)
        self.text_jyff.delete(1.0,100.0) #小数点左边行号，从1开始
        self.text_czbz.delete(1.0,100.0) #小数点右边列号，从0开始
        self.listbox_example.delete(0,9)
        self.text_addexample.delete(1.0,100.0)
        self.play_count = 0 
        
        if designate==None:
            n = self.row_number-1 # should be list length
            self.current_index = random.randrange(0,n)
        else:
            self.current_index = self.list_word.index(designate)
            

        #self.word_entry = self.list_word[self.current_index]
        #print(self.word_entry)
        if self.list_ipu[self.current_index] == 'improtant':    
            #OK 变色复用模块
            self.entry_word['bg'] = COLOR_BG_I 
            self.entry_word['fg'] = COLOR_FG_I
        elif self.list_ipu[self.current_index] == 'practice':
            self.entry_word['bg'] = COLOR_BG_P 
            self.entry_word['fg'] = COLOR_FG_P
        elif self.list_ipu[self.current_index] == 'unnecessary':
            self.entry_word['bg'] = COLOR_BG_U 
            self.entry_word['fg'] = COLOR_FG_U
        else:
            self.entry_word['bg'] = COLOR_BG_N
            self.entry_word['fg'] = COLOR_FG_N
        
        print(self.list_word[self.current_index])
        self.testedlog.append(self.list_word[self.current_index])
        self.process_repeat()
        
        
    def set_ipu(self,ipu):
        if ipu == 'i':
            self.entry_word['bg'] = COLOR_BG_I
            self.entry_word['fg'] = COLOR_FG_I
            self.add('ipu','important')
        elif ipu == 'p':
            self.entry_word['bg'] = COLOR_BG_P 
            self.entry_word['fg'] = COLOR_FG_P
            self.add('ipu','practice')
        elif ipu == 'u':
            self.entry_word['bg'] = COLOR_BG_U 
            self.entry_word['fg'] = COLOR_FG_U
            self.add('ipu','unnecessary')
        
    
    def process_next(self):
        '''
        现在确定的逻辑：
            1. 播放音频，不显示任何内容
            2.1 空内容按enter，显示释义
            2.2 输入正确后显示记忆方法、备注和例句
            2.2 输入错误进行反馈
            3 空内容enter，且已经显示释义...
            ...

        Returns
        -------
        None.

        '''
        #TODO 记忆方法和词组备注在答对后再显示
        
        #print("next")
        #输入正确，且没show
        if self.word_entry.get() == self.list_word[self.current_index] and not self.is_shown:
            self.process_show()
        #输入正确，且已经show
        elif self.word_entry.get() == self.list_word[self.current_index] and self.is_shown:
            self.nextword()
        #输入错误，且没show
        elif not self.is_shown and self.word_entry.get() != '':
            #self.process_show()
            self.entry_word['fg'] = COLOR_WRONG_WORD
        #输入错误，且已经show
        elif self.is_shown and self.word_entry.get() != '':
            self.entry_word['fg'] = COLOR_WRONG_WORD
        #没有输入，且没show
        elif self.word_entry.get() == '' and not self.is_shown:
            self.process_show()
        #没有输入且已经show
        elif self.word_entry.get() == '' and self.is_shown:
            self.entry_word['fg'] = COLOR_PASS
            self.word_entry.set(self.list_word[self.current_index])

            
    def process_enter(self,event):
        self.process_next()
        
    def process_prev(self):
        pass
        if len(self.testedlog)>1:
            self.nextword(self.testedlog[-2])
            self.process_show()
                
    def process_repeat(self):
#        path=os.path.join(PATH_AUDIO,"%s_%d.mp3"%(self.list_word[self.current_index],self.play_count%2+1))
        path=PATH_AUDIO+"\%s_%d.mp3"%(self.list_word[self.current_index],self.play_count%2+1)
        # OK 在播放声音前确认是否存在
        # print(path)
        if os.path.exists(path):
            self.play_sound(path)
        else:
            # TODO 未找到放一段提示音效？
            print('音频未找到')

        self.play_count += 1
        
    def process_show(self):
        #print(self.is_shown)
        if not self.is_shown:
            #print('show',self.list_word[self.current_index])
            
            #self.word_entry.set(self.list_word[self.current_index])
            
            temp1 = len(self.list_paraphrase[self.current_index])
            for i in range(temp1):
                #对人名的特殊处理
                #if self.list_word[self.current_index][1:] in self.list_paraphrase[self.current_index][i]:
                if '人名' in self.list_paraphrase[self.current_index][i]:
                    self.listbox_paraphrase.insert(i,'n.人名')
                else:
                    self.listbox_paraphrase.insert(i,self.list_paraphrase[self.current_index][i])
#            self.text_jyff['text'] = self.list_jyff[self.current_index] #TODO buduide
#            self.text_czbz['text'] = self.list_czbz[self.current_index]
            self.text_jyff.insert(INSERT,str(self.list_jyff[self.current_index]))
            self.text_czbz.insert(INSERT,str(self.list_czbz[self.current_index]))
            temp2 = len(self.list_example[self.current_index])
            for i in range(temp2):
                self.listbox_example.insert(i,self.list_example[self.current_index][i])
            self.is_shown = True
        else:
            self.word_entry.set(self.list_word[self.current_index])
            
        #print(self.text_czbz.get(1.0,100.0)[:-1])
            
    def process_change(self,to,content):
        index=self.current_index
        log=[index,to,content]
        self.changelog.append(log)
            
    def open_url(self,hint):
        if hint == 'bing':
            webbrowser.open('https://cn.bing.com/dict/search?q=%s'%self.list_word[self.current_index]) #TODO
        elif hint == 'youdao':
            webbrowser.open('https://www.dict.youdao.com') #TODO
    def process_save(self):
        self.save_excel()
    
    #将改动加入changelog
    def add(self,to,content):
        index=self.current_index+2
        if to == 'ipu':
            log=[index,1,content] #row, column, content
            self.changelog.append(log)
            self.list_ipu[self.current_index] = content
        elif to == 'jyff':
            log=[index,4,content] #row, column, content
            self.changelog.append(log)
            self.list_jyff[self.current_index] = content
        elif to == 'czbz':
            log=[index,5,content] #row, column, content
            self.changelog.append(log)
            self.list_czbz[self.current_index] = content
        elif to == 'example':
            if(content!=""):
                #OK special treat
                #还要分有没有之前例句的情况，例句是list in list
                if self.list_example[self.current_index]==[]:
                    #之前没有例句
                    new='="'+content+'"'
                elif self.list_example[self.current_index]!=[]:
                    #之前有例句，
                    if(self.list_example[self.current_index][0]=="="):
                        #且以=开头
                        new = self.list_example[self.current_index]+"&CHAR(10)&"+'"'+content+'"'
                    else:
                        #且不以=开头，就说明是录入单词的时候直接添加的
                        print("之前存在未格式化的例句！")
                        new = '="'+self.list_example[self.current_index]+'"'+"&CHAR(10)&"+'"'+content+'"'
                
                log=[index,6,new]
                print(log)
                self.changelog.append(log)
                
                #OK 清空输入内容并加入上方例句表
                #self.listbox_example.insert(i,self.list_example[self.current_index][i])
                i = len(self.list_example[self.current_index])
                new_content = []
                new_content = content.split("\n")
                print(new_content)
                for content in new_content:
                    i+=1
                    self.listbox_example.insert(i+1,content)
                    self.text_addexample.delete(1.0,100.0)
                
                self.list_example[self.current_index] = new
                
        print(self.changelog)
        
    def save_excel(self):
        workbook = load_workbook(PATH_SHANBAY)#找到excel文件
        sheet = workbook["Sheet1"]#找到当前表格
        if(self.changelog!=[]):
            for item in self.changelog:
                sheet.cell(row=item[0],column=item[1]).value = item[2]
            try:
                #save override
                workbook.save(PATH_SHANBAY)
                print('saved')
            except:
                #save another
                workbook.save(PATH_SHANBAY_BACKUP)
                print('saved to backup location')
            
            self.changelog = []

    def closewin(self):
        if tkinter.messagebox.askyesno("退出","你确定要退出吗？"):
            if(self.changelog!=[]):
                if tkinter.messagebox.askyesno("未保存","有未保存的改动，是否保存？"):
                    self.save_excel()
            
            # TODO 保存log
            # with open 
            print("exit")
            self.win.destroy()
            
        # else:
        #     pass

def main():
    #window = Tk()
    T = Tester()
    
    #window.mainloop()
    
main()
