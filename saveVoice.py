# -*- coding: utf-8 -*-
"""
Created on Mon Oct 22 12:26:16 2018

@author: HP
"""

#从有道词典获取单词发音并储存到本地

import urllib.request
from os.path import join

ROOT = r'voices'
def saveVoice(word,typ=0):
    if typ == 1:#英声
        url = 'http://dict.youdao.com/dictvoice?audio=%s&type=1'%word
        urllib.request.urlretrieve(url,join(ROOT,"%s_1.mp3"%word))
    elif typ == 2:#美声
        url = 'http://dict.youdao.com/dictvoice?audio=%s&type=2'%word
        urllib.request.urlretrieve(url,join(ROOT,"%s_2.mp3"%word))
    elif typ == 0:#英美声都下载
        url = 'http://dict.youdao.com/dictvoice?audio=%s&type=1'%word
        urllib.request.urlretrieve(url,join(ROOT,"%s_1.mp3"%word))
        url = 'http://dict.youdao.com/dictvoice?audio=%s&type=2'%word
        urllib.request.urlretrieve(url,join(ROOT,"%s_2.mp3"%word))
    print(word,"done")
    
def temp():
    from openpyxl import load_workbook
        #print("正在打开excel文件...")
    workbook = load_workbook(u'shanbay1.xlsx')#找到excel文件
    sheet = workbook.get_sheet_by_name("Sheet1")#找到当前表格
    row_number = sheet.max_row
    word_list =[]
    for i in range(885,row_number+1):
        word = (str(sheet.cell(row=i,column=2).value))
        print(word)
        saveVoice(word)
        print(word,"done",i)