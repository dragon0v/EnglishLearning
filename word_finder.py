# -*- coding: utf-8 -*-
"""
Created on Tue Sep 15 18:02:56 2020

@author: ToxicNeoBanana
"""

#word searcher
# done 加入例句后会有bug --已增加do_clear中的列长度，应该ok
# TODO 例句只能加一条
# TODO undo last changes, 现在只是不对数据库产生影响，软件层面还没有undo一些添加的效果
'''
  File "F:/OneDrive/word_finder.py", line 349, in add
    if(self.list_example_raw[self.current_index][0]=="="):
TypeError: 'NoneType' object is not subscriptable
'''

from tkinter import *
import tkinter.messagebox
import tkinter.font as tf
import time
import datetime
import os
import random
import webbrowser
from openpyxl import load_workbook
import pickle
from collections import Counter


#以下是全局变量
print("Hello,",os.environ['USERNAME']) #可能不成功因为与环境变量有关
USERNAME = os.environ['USERNAME']
if USERNAME=="ToxicNeoBanana":
    PATH_SHANBAY = r"F:\OneDrive\shanbay1.xlsx"
    PATH_SHANBAY_BACKUP = os.path.join(os.path.expanduser('~'),"Desktop","shanbay1-%d.xlsx"%int(time.time()))
    PATH_AUDIO = r"D:\python\英语学习\voices"
    PATH_SEARCHPY = r"D:\python\英语学习\批量搜词写入excel - 自动换行下载音频.py"
    PATH_SHANBAY_FREQ = r"F:\OneDrive\shanbay_freq.pkl"

elif USERNAME=="NeoBanana":
    PATH_SHANBAY = r"C:\Users\NeoBanana\OneDrive\shanbay1.xlsx"
    PATH_SHANBAY_BACKUP = os.path.join(os.path.expanduser('~'),"Desktop","shanbay1-%d.xlsx"%int(time.time()))
    PATH_AUDIO = r"D:\python\英语学习\voices"
    PATH_SEARCHPY = r"D:\python\英语学习\批量搜词写入excel - 自动换行下载音频.py"

COLOR_BG_I = '#ffafaf'
COLOR_FG_I = '#ff2727'
COLOR_BG_P = '#c0e399'
COLOR_FG_P = '#548235'
COLOR_BG_U = '#acb9ca'
COLOR_FG_U = '#333f4f'
COLOR_BG_N = '#F7F762'
COLOR_FG_N = 'black'
COLOR_BG_G = '#BDD7EE'  # GRE常考，默认为practice的水平，只有真题见到才打这个标签
COLOR_FG_G = '#0070C0'
COLOR_WRONG_WORD = 'red'
COLOR_PASS = '#444' #回车过去的单词颜色

URL_MERRIAM_WEBSTER = r'https://www.merriam-webster.com/dictionary/' #在最后加要查的词

print("Last modified:",time.ctime(os.path.getmtime(PATH_SHANBAY)))
class Finder():
    def __init__(self):
        self.win=Tk()
        self.win.geometry('480x512+800+380')
        self.win.title("word_finder")
        
        with open(PATH_SHANBAY_FREQ,'rb') as f:
            self.shanbay_freq = pickle.load(f)

        self.place_widgets()
        
        self.load_excel()
        
        self.changelog=[] # save all changes
        self.can_add = False
        
        #self.nextword()
        
        #此处可以用于debug print
        self.is_shown = False
        
        
#        self.win.bind_all("<KeyPress-P>",self.save_excel)
        self.win.protocol("WM_DELETE_WINDOW",self.closewin)#按关闭键执行self.closewin
        
        self.win.mainloop()
    
    def place_widgets(self):
        self.canvas = Canvas(self.win,width=900,height=600,background='lightgrey')
        self.canvas.pack()
        self.canvas.bind_all("<Control-KeyPress-S>",self.saaave)
        
        self.word_entry = StringVar()
        ft = tf.Font(family='mv boli', size='14')
        self.entry_word = Entry(self.canvas,bg='yellow',textvariable=self.word_entry,font=ft)
        self.entry_word.place(x=10,y=10,width=160,height=26,anchor=NW)
        self.entry_word.bind("<Return>",self.process_enter)        
        self.button_find = Button(self.canvas,text="find",command=lambda:self.process_find())
        self.button_find.place(x=180,y=10,width=60,height=50,anchor=NW)
        self.button_add = Button(self.canvas,text="add",command=lambda:self.process_add())
        self.button_add.place(x=250,y=10,width=60,height=50,anchor=NW)
        self.button_clear = Button(self.canvas,text="clear",command=lambda:self.process_clear())
        self.button_clear.place(x=320,y=10,width=60,height=50,anchor=NW)
        self.button_baidu = Button(self.canvas,text="bing",command=lambda:self.open_url('bing'))
        self.button_baidu.place(x=390,y=10,width=60,height=24,anchor=NW)
        self.button_youdao = Button(self.canvas,text="Cambridge",command=lambda:self.open_url('youdao'))
        self.button_youdao.place(x=390,y=36,width=60,height=24,anchor=NW)
        self.button_save = Button(self.canvas,text="存",command=lambda:self.process_save())
        self.button_save.place(x=460,y=10,width=14,height=14,anchor=NW)
#        self.button_search = Button(self.canvas,text="搜",command=lambda:self.process_search())
#        self.button_search.place(x=460,y=28,width=14,height=14,anchor=NW)
        self.button_search = Button(self.canvas,text="撤",command=lambda:self.process_undo())
        self.button_search.place(x=460,y=28,width=14,height=14,anchor=NW)
        self.button_open = Button(self.canvas,text="开",command=lambda:self.process_open())
        self.button_open.place(x=460,y=46,width=14,height=14,anchor=NW)
        
        self.button_ipu_i = Button(self.canvas,text="",command=lambda:self.set_ipu('i'),background=COLOR_BG_I)
        self.button_ipu_i.place(x=10,y=40,width=50,height=20,anchor=NW)
        self.button_ipu_p = Button(self.canvas,text="",command=lambda:self.set_ipu('p'),background=COLOR_BG_P)
        self.button_ipu_p.place(x=65,y=40,width=50,height=20,anchor=NW)
        self.button_ipu_u = Button(self.canvas,text="",command=lambda:self.set_ipu('u'),background=COLOR_BG_U)
        self.button_ipu_u.place(x=120,y=40,width=20,height=20,anchor=NW)
        self.button_ipu_u = Button(self.canvas,text="",command=lambda:self.set_ipu('g'),background=COLOR_BG_G)
        self.button_ipu_u.place(x=150,y=40,width=20,height=20,anchor=NW)
        
        self.freq_entry = IntVar()
        self.button_freq_down = Button(self.canvas,text="-",command=lambda:self.process_add_freq(-1))
        self.button_freq_down.place(x=120,y=70,width=20,height=20,anchor=NW)
        self.button_freq_up = Button(self.canvas,text="+",command=lambda:self.process_add_freq(1))
        self.button_freq_up.place(x=150,y=70,width=20,height=20,anchor=NW)
        self.entry_freq = Entry(self.canvas,textvariable=self.freq_entry)
        self.entry_freq.place(x=180,y=70,width=20)
        
        self.label_paraphrase = Label(self.canvas,text="paraphrase")
        self.label_paraphrase.place(x=10,y=70)
        self.listbox_paraphrase = Listbox(self.canvas)
        self.listbox_paraphrase.place(x=10,y=90,width=230,height=200,anchor=NW)
        self.label_jyff = Label(self.canvas,text="记忆方法")
        self.label_jyff.place(x=250,y=70)
        self.button_addjyff = Button(self.canvas,text="add",command=lambda:self.add('jyff',self.text_jyff.get(1.0,100.0)[:-1])) #:-1去掉最后的换行
        self.button_addjyff.place(x=410,y=72,width=40,height=20,anchor=NW)
        self.text_jyff = Text(self.canvas)
        self.text_jyff.place(x=250,y=90,width=200,height=90,anchor=NW)
        self.label_czbz = Label(self.canvas,text="词组备注")
        self.label_czbz.place(x=250,y=180)
        self.button_addczbz = Button(self.canvas,text="add",command=lambda:self.add('czbz',self.text_czbz.get(1.0,100.0)[:-1]))
        self.button_addczbz.place(x=410,y=182,width=40,height=20,anchor=NW)
        self.text_czbz = Text(self.canvas)
        self.text_czbz.place(x=250,y=200,width=200,height=90,anchor=NW)
        
        
        self.label_example = Label(self.canvas,text="example")
        self.label_example.place(x=10,y=300)
        self.listbox_example = Listbox(self.canvas)
        self.listbox_example.place(x=10,y=320,width=840,height=104,anchor=NW)
        
        self.text_addexample = Text(self.canvas)
        self.text_addexample.place(x=10,y=434,width=400,height=72,anchor=NW)
        self.button_addexample = Button(self.canvas,text="add",command=lambda:self.add('example',self.text_addexample.get(1.0,100.0)[:-1]))
        self.button_addexample.place(x=410,y=434,width=40,height=20,anchor=NW)
        
    def load_excel(self):
        workbook = load_workbook(PATH_SHANBAY)#找到excel文件
        sheet = workbook["Sheet1"]#找到当前表格
        self.row_number = sheet.max_row
        self.current_index = self.row_number
        print("Total entries:",self.row_number)
        
        self.list_ipu = []
        self.list_word = []
        list_paraphrase_raw = [] #need further process
        self.list_jyff = []
        self.list_czbz = []
        list_example_raw = [] #
        
        for i in range(2,self.row_number+1):
            self.list_ipu.append(sheet.cell(row=i,column=1).value)
            self.list_word.append(sheet.cell(row=i,column=2).value)
            list_paraphrase_raw.append(sheet.cell(row=i,column=3).value)
            self.list_jyff.append(sheet.cell(row=i,column=4).value)
            self.list_czbz.append(sheet.cell(row=i,column=5).value)
            list_example_raw.append(sheet.cell(row=i,column=6).value)
        
        #print(self.list_czbz[:10])
        
        self.list_paraphrase = self.parse_paraphrase(list_paraphrase_raw) # list in list
        self.list_example = self.parse_paraphrase(list_example_raw) # 
        #print(self.list_paraphrase[:10])
        
        # 2021-12-9 更新，对于之前存在的例句，添加时直接在其后追加？
        self.list_example_raw = list_example_raw
            
    def parse_paraphrase(self,list_paraphrase):
        #="n. 耙"&CHAR(10)&"vt. 耙地；使苦恼"&CHAR(10)&"vi. 被耙松"&CHAR(10)&"n. (Harrow)人名；(英)哈罗"
        after = []
        for each in list_paraphrase:
            if each == None:
                after.append([])
            else:
                p1 = each[2:-1]
                p2 = p1.split('"&CHAR(10)&"')
                after.append(p2) # list in list
        return after
    
    def set_ipu(self,ipu):
        if ipu == 'i':
            self.entry_word['bg'] = COLOR_BG_I
            self.entry_word['fg'] = COLOR_FG_I
            self.add('ipu','important')
        elif ipu == 'p':
            self.entry_word['bg'] = COLOR_BG_P
            self.entry_word['fg'] = COLOR_FG_P
            self.add('ipu','practice')
        elif ipu == 'u':
            self.entry_word['bg'] = COLOR_BG_U
            self.entry_word['fg'] = COLOR_FG_U
            self.add('ipu','unnecessary')
        elif ipu == 'g':
            self.entry_word['bg'] = COLOR_BG_G
            self.entry_word['fg'] = COLOR_FG_G
            self.add('ipu','GRE常考')
        
    
    def process_find(self):
        self.do_clear()
        #输入的单词已经存在
        if self.word_entry.get() in self.list_word:
            print(f"{self.word_entry.get()} 已经存在")
            self.is_shown = True
            #TODO +2
            self.current_index = self.list_word.index(self.word_entry.get())
            #显示释义等内容
            self.process_show()
            if self.list_ipu[self.current_index] == 'important':    
                #OK 变色复用模块
                self.entry_word['bg'] = COLOR_BG_I 
                self.entry_word['fg'] = COLOR_FG_I
            elif self.list_ipu[self.current_index] == 'practice':
                self.entry_word['bg'] = COLOR_BG_P 
                self.entry_word['fg'] = COLOR_FG_P
            elif self.list_ipu[self.current_index] == 'unnecessary':
                self.entry_word['bg'] = COLOR_BG_U 
                self.entry_word['fg'] = COLOR_FG_U
            elif self.list_ipu[self.current_index] == 'GRE常考':
                self.entry_word['bg'] = COLOR_BG_G
                self.entry_word['fg'] = COLOR_FG_G
            else:
                self.entry_word['bg'] = COLOR_BG_N
                self.entry_word['fg'] = COLOR_FG_N
        else:
            self.entry_word['fg'] = COLOR_WRONG_WORD
            print("不存在，是否添加")
            
            
    def process_enter(self,event):
        if self.word_entry.get() == "":
            return
        if self.is_shown and self.word_entry.get() in self.list_word:
            self.process_clear()
            self.is_shown = False
        else:
            self.process_find()
    
    def process_undo(self):
        if len(self.changelog)>0:
            self.changelog.pop()
            print(self.changelog)
        else:
            print('no changes')
    
    #add有find的所有功能，并有添加单词的作用
    def process_add(self):
        #OK reset模块封装
        self.do_clear()
        if self.word_entry.get() in self.list_word:
            print("已经存在，无需添加")
            self.is_shown = True
            self.current_index = self.list_word.index(self.word_entry.get())
            self.process_show()
            if self.list_ipu[self.current_index] == 'improtant':    
                #TODO 变色功能封装
                self.entry_word['bg'] = COLOR_BG_I 
                self.entry_word['fg'] = COLOR_FG_I
            elif self.list_ipu[self.current_index] == 'practice':
                self.entry_word['bg'] = COLOR_BG_P 
                self.entry_word['fg'] = COLOR_FG_P
            elif self.list_ipu[self.current_index] == 'unnecessary':
                self.entry_word['bg'] = COLOR_BG_U 
                self.entry_word['fg'] = COLOR_FG_U
            else:
                self.entry_word['bg'] = COLOR_BG_N
                self.entry_word['fg'] = COLOR_FG_N
        else:
            #添加部分
            self.row_number += 1
            self.current_index = self.row_number - 2 #legacy
            log=[self.current_index+2,2,self.word_entry.get()]
            self.changelog.append(log)
            print(self.changelog)
            print("已添加")
            self.list_word.append(self.word_entry.get())
            self.list_ipu.append("")
            self.list_czbz.append("")
            self.list_jyff.append("是你刚添加的词")
            self.list_example.append([]) #新添加的词
            self.list_paraphrase.append("")
            
            
    def process_clear(self):
        self.is_shown = False
        self.do_clear()
        self.word_entry.set("")
        
    def do_clear(self):
        self.listbox_paraphrase.delete(0,900)
        self.text_jyff.delete(1.0,1000.0) #小数点左边行号，从1开始
        self.text_czbz.delete(1.0,1000.0) #小数点右边列号，从0开始
        self.listbox_example.delete(0,900)
        self.text_addexample.delete(1.0,5000.0)
        self.entry_word['bg'] = COLOR_BG_N
        self.entry_word['fg'] = COLOR_FG_N
        self.play_count = 0
        
    def process_show(self):
        temp1 = len(self.list_paraphrase[self.current_index])
        for i in range(temp1):
            self.listbox_paraphrase.insert(i,self.list_paraphrase[self.current_index][i])
        self.text_jyff.insert(INSERT,str(self.list_jyff[self.current_index]))
        self.text_czbz.insert(INSERT,str(self.list_czbz[self.current_index]))
        temp2 = len(self.list_example[self.current_index])
        for i in range(temp2):
            self.listbox_example.insert(i,self.list_example[self.current_index][i])
        
        self.update_freq()
        
#    def process_change(self,to,content):
#        index=self.current_index
#        log=[index,to,content]
#        self.changelog.append(log)
            
    def open_url(self,hint):
        if hint == 'bing':
            webbrowser.open('https://cn.bing.com/dict/search?q=%s'%self.word_entry.get())
        elif hint == 'youdao':
#            webbrowser.open('https://www.dict.youdao.com') #TODO
            webbrowser.open('https://dictionary.cambridge.org/dictionary/english-chinese-simplified/%s'%self.word_entry.get())
    def process_save(self):
        self.save_excel()
    
    def process_search(self):
        os.startfile(PATH_SEARCHPY)
        
    def process_open(self):
        os.startfile(PATH_SHANBAY)
    
    def update_freq(self):
        # 使stringvar的值永远等于Counter的值
        self.freq_entry.set(self.shanbay_freq[self.word_entry.get()])
    
    def process_add_freq(self,v):
        self.shanbay_freq[self.word_entry.get()] += v
        self.update_freq()
#        self.freq_entry = str(self.shanbay_freq[self.word_entry.get()])
        print(self.word_entry.get(),self.shanbay_freq[self.word_entry.get()])

        
        
    #将改动加入changelog
    def add(self,to,content):
        #index=self.current_index+2
        if to == 'ipu':
            log=[self.current_index+2,1,content] #row, column, content
            self.changelog.append(log)
            self.list_ipu[self.current_index] = content
        elif to == 'jyff':
            log=[self.current_index+2,4,content] #row, column, content
            self.changelog.append(log)
            self.list_jyff[self.current_index] = content
        elif to == 'czbz':
            log=[self.current_index+2,5,content] #row, column, content
            self.changelog.append(log)
            self.list_czbz[self.current_index] = content
        elif to == 'example':
            if(content!=""):
                #OK special treat
                #还要分有没有之前例句的情况，例句是list in list
                if self.list_example[self.current_index]==[]:
                    #之前没有例句
                    new='="'+content+'"'
                elif self.list_example[self.current_index]!=[]:
                    #之前有例句，
                    if(self.list_example_raw[self.current_index][0]=="="):
                        #且以=开头
                        new = self.list_example_raw[self.current_index]+"&CHAR(10)&"+'"'+content+'"'
                        new = '''="12 months constitute a year."&CHAR(10)&"7 days constitute a week."&CHAR(10)&"In the following sections,we describe five different methods, which constitute potential solutions to the FN problem."'''
                        print(new)
                    else:
                        #且不以=开头，就说明是录入单词的时候直接添加的
                        print("之前存在未格式化的例句！")
                        new = '="'+self.list_example_raw[self.current_index]+'"'+"&CHAR(10)&"+'"'+content+'"'
                
                log=[self.current_index+2,6,new]
                self.changelog.append(log)
                
                #OK 清空输入内容并加入上方例句表
                #self.listbox_example.insert(i,self.list_example[self.current_index][i])
                i = len(self.list_example[self.current_index])
                new_content = []
                new_content = content.split("\n")
                print(new_content)
                for content in new_content:
                    i+=1
                    self.listbox_example.insert(i+1,content)
                    self.text_addexample.delete(1.0,100.0)
                
                self.list_example[self.current_index] = new
                
        print(self.changelog)
        
    def saaave(self,event):
        self.save_excel()
    
    def save_excel(self):
        with open(PATH_SHANBAY_FREQ,'wb') as f:
            pickle.dump(self.shanbay_freq,f)
        
        workbook = load_workbook(PATH_SHANBAY)#找到excel文件
        sheet = workbook["Sheet1"]#找到当前表格
        if(self.changelog!=[]):
            for item in self.changelog:
                sheet.cell(row=item[0],column=item[1]).value = item[2]
            try:
                #save override
                workbook.save(PATH_SHANBAY)
                print('saved')
            except:
                #save another
                workbook.save(PATH_SHANBAY_BACKUP)
                print('saved to backup location')
            
            self.changelog = []
        else:
            print('no unsaved change')
        
    def closewin(self):
        if tkinter.messagebox.askyesno("退出","你确定要退出吗？"):
            if(self.changelog!=[]):
                if tkinter.messagebox.askyesno("未保存","有未保存的改动，是否保存？"):
                    self.save_excel()
            with open(PATH_SHANBAY_FREQ,'wb') as f:
                pickle.dump(self.shanbay_freq,f)
            print("exit")
            self.win.destroy()
        else:
            pass

def main():
    F = Finder()
    
main()
