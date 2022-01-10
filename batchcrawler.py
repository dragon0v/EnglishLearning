# -*- coding: utf-8 -*-
"""
Created on Mon Aug 13 21:38:54 2018

@author: HP
"""

#shanbay批量搜词写入shanbay.xlsx
#2018-8-13 21:39:46
#加入try结构 细化

#v1.1 2018-10-22 加入下载有道语音功能
#v1.2 2019-1-9 去掉了try的结构-什么垃圾以后不用了-等等要不干脆重写吧
#2019-1-23 改成绝对路径
#2019-12-29 加入sleep防反爬虫

#打开部分
from openpyxl import load_workbook
import time
import random

flag = True#未出错 = true
try:
    print("正在打开excel文件...")
    workbook = load_workbook(r"F:\OneDrive\shanbay1.xlsx")#找到excel文件
    sheet = workbook.get_sheet_by_name("Sheet1")#找到当前表格
    row_number = sheet.max_row
except:
    print("打开失败")
    flag = False
else:
    print("打开成功!\n")


#读取部分
from crawler import *
from saveVoice import saveVoice
import urllib.request


if flag == True:
    try:
        print("正在读取excel文件...")
        reslist = [] #result list
        word = ""
        youbian = ""
        collected=0#获取20个词之后sleep一段时间，并保存
        for i in range(1,row_number+1):
            word = str(sheet.cell(row=i,column=2).value)            
            youbian = str(sheet.cell(row=i,column=3).value)
            print("正在获取翻译"+str(i)+"/"+str(row_number),end='\t')
            if word == "单词" or word == "" or youbian != "None":
                reslist.append("")
                print(word,end='\t\t')
                print(youbian[:40])
            else:
                #TODO saveVoice可能出现迷之错误
                saveVoice(word)
                fetch = get_shiyi(word)
                string = "="
                for each in fetch:
                    string += ('"'+each+'"&CHAR(10)&')
                reslist.append(string[:-10])
                if collected==20:
                    
                    collected=0
                    print('sleeeeping ')
                    for o in range(1,16):
                        time.sleep(1)#每查20次sleep 15秒防反爬虫
                        print(o)
                    print('waking')
                else:
                    collected+=1
                
        #assert len(reslist) == row_number
        
    except:
        print("文件读取出现问题或者断网了")

        flag = False
    else:
        if reslist.count("") == row_number :
            print("无新词")
            flag = False
        else:
            #print("读取成功!")
            pass
    finally:
        #del sheet
        #del workbook
        pass
    
        
#写入部分
if flag == True:

    print("正在写入...")
    t = 0
    for j in sheet["C"]:
        if reslist[t] != "":
            j.value = reslist[t]
        t += 1
        
    workbook.save(r"F:\OneDrive\shanbay1.xlsx")

#print('all done')
input('all done')








